import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill


def export_dcf_report(
    ticker,

    revenue_cagr,
    average_margin,

    revenue,
    ebit,
    ebit_margin,

    revenue_forecast,
    ebit_forecast,
    nopat_forecast,
    fcff_forecast,

    forecast_growth_rate,

    tax_rate,
    depreciation_percent,
    capex_percent,

    risk_free_rate,
    equity_risk_premium,
    beta,
    cost_of_debt,
    wacc,

    terminal_growth_rate,

    terminal_value,
    pv_terminal_value,

    enterprise_value,
    equity_value,
    fair_value_per_share,

    sensitivity_table,

    current_price,
    upside_downside,
    recommendation
):

    # Create file name
    file_name = f"DCF_Report_{ticker}.xlsx"

    # --------------------------------------------------
    # DASHBOARD SHEET
    # --------------------------------------------------

    dashboard_df = pd.DataFrame({

        "Metric": [

            "Current Market Price",

            "DCF Fair Value",

            "Upside / Downside",

            "Recommendation",

            "Enterprise Value",

            "Equity Value"

        ],

        "Value": [

            current_price,

            fair_value_per_share,

            upside_downside,

            recommendation,

            enterprise_value,

            equity_value
        ]
    })

    # --------------------------------------------------
    # ASSUMPTIONS SHEET
    # --------------------------------------------------

    assumptions_df = pd.DataFrame({

        "Assumption": [

            "Revenue CAGR",

            "Forecast Growth Rate",

            "Average EBIT Margin",

            "Tax Rate",

            "Depreciation %",

            "CapEx %",

            "Risk Free Rate",

            "Equity Risk Premium",

            "Beta",

            "Cost of Debt",

            "WACC",

            "Terminal Growth Rate"

        ],

        "Value": [

            revenue_cagr,

            forecast_growth_rate,

            average_margin,

            tax_rate * 100,

            depreciation_percent * 100,

            capex_percent * 100,

            risk_free_rate * 100,

            equity_risk_premium * 100,

            beta,

            cost_of_debt * 100,

            wacc * 100,

            terminal_growth_rate * 100
        ]
    })

    # --------------------------------------------------
    # HISTORICAL ANALYSIS SHEET
    # --------------------------------------------------

    historical_df = pd.DataFrame({

        "Revenue": revenue,

        "EBIT": ebit,

        "EBIT Margin (%)": ebit_margin
    })

    # --------------------------------------------------
    # FORECAST SHEET
    # --------------------------------------------------

    forecast_df = pd.DataFrame({

        "Year": [

            "Year 1",

            "Year 2",

            "Year 3",

            "Year 4",

            "Year 5"
        ],

        "Revenue": revenue_forecast,

        "EBIT": ebit_forecast,

        "NOPAT": nopat_forecast,

        "FCFF": fcff_forecast
    })

    # --------------------------------------------------
    # DCF VALUATION SHEET
    # --------------------------------------------------

    dcf_df = pd.DataFrame({

        "Metric": [

            "Terminal Value",

            "PV Terminal Value",

            "Enterprise Value",

            "Equity Value",

            "Fair Value Per Share"
        ],

        "Value": [

            terminal_value,

            pv_terminal_value,

            enterprise_value,

            equity_value,

            fair_value_per_share
        ]
    })

    # --------------------------------------------------
    # WRITE EXCEL FILE
    # --------------------------------------------------

    with pd.ExcelWriter(
        file_name,
        engine="openpyxl"
    ) as writer:

        dashboard_df.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False
        )

        assumptions_df.to_excel(
            writer,
            sheet_name="Assumptions",
            index=False
        )

        historical_df.to_excel(
            writer,
            sheet_name="Historical Analysis"
        )

        forecast_df.to_excel(
            writer,
            sheet_name="Forecast",
            index=False
        )

        dcf_df.to_excel(
            writer,
            sheet_name="DCF Valuation",
            index=False
        )

        sensitivity_table.to_excel(
            writer,
            sheet_name="Sensitivity"
        )

    # --------------------------------------------------
    # FORMAT WORKBOOK
    # --------------------------------------------------

    workbook = load_workbook(file_name)

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    for sheet in workbook.worksheets:

        # Format headers

        for cell in sheet[1]:

            cell.fill = header_fill

            cell.font = header_font

        # Freeze first row

        sheet.freeze_panes = "A2"

        # Auto-size columns

        for column in sheet.columns:

            max_length = 0

            column_letter = column[0].column_letter

            for cell in column:

                try:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

                except:

                    pass

            sheet.column_dimensions[
                column_letter
            ].width = max_length + 2

    workbook.save(file_name)

    return file_name